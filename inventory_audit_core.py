"""Core inventory parsing and audit report generation for Inventory Audit."""

from __future__ import annotations

import argparse
import csv
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Sequence

from inventory_audit_requirements import ART_FIELDS, ENDPOINT_ORDER, get_art_requirements, get_media_requirements


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif"}
MEDIA_EXTENSIONS = {".mov", ".vtt", ".srt"}
ART_RE = re.compile(
    r"(?:^|[_-])(?P<kind>ca|bg|tt)[_-](?P<ratio>\d+x\d+)[_-](?P<width>\d{3,5})x(?P<height>\d{3,5})(?=\.[a-z0-9]+$)",
    re.IGNORECASE,
)
SEASON_RE = re.compile(r"^s(?P<num>\d{1,3})$", re.IGNORECASE)
EPISODE_RE = re.compile(r"^e(?P<num>\d{1,3})$", re.IGNORECASE)
UUID_RE = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.IGNORECASE)


class InventoryAuditError(ValueError):
    """Raised when an inventory file cannot be audited."""


@dataclass(frozen=True)
class InventoryItem:
    bucket: str
    key: str
    size_bytes: int = 0
    last_modified: str = ""
    s3_uri: str = ""

    @property
    def filename(self) -> str:
        return Path(self.key).name

    @property
    def suffix(self) -> str:
        return Path(self.key).suffix.lower()


@dataclass
class AuditEntity:
    content_type: str
    name: str
    title: str
    sku: str
    bucket: str
    path_key: str
    season: str = ""
    episode: str = ""
    files: list[InventoryItem] = field(default_factory=list)

    @property
    def s3_path(self) -> str:
        return f"s3://{self.bucket}/{self.path_key}" if self.bucket else self.path_key


@dataclass(frozen=True)
class AuditResult:
    csv_path: Path
    xlsx_path: Path
    rows: list[OrderedDict[str, str]]
    source_file_count: int
    inventory_csv_path: Path | None = None

    @property
    def output_path(self) -> Path:
        return self.csv_path

    @property
    def output_paths(self) -> tuple[Path, Path]:
        return self.csv_path, self.xlsx_path

    @property
    def entity_count(self) -> int:
        return len(self.rows)


def read_inventory(path: Path | str) -> tuple[str, list[InventoryItem]]:
    """Read CSV/XLSX S3 inventory exports into normalized inventory items."""

    source = Path(path)
    if not source.exists():
        raise InventoryAuditError(f"Inventory file not found: {source}")

    suffix = source.suffix.lower()
    if suffix == ".csv":
        inventory_uri, rows = _read_csv_inventory(source)
    elif suffix in {".xlsx", ".xlsm"}:
        inventory_uri, rows = _read_xlsx_inventory(source)
    elif suffix == ".xls":
        raise InventoryAuditError("Legacy .xls files are not supported in v1. Save as .xlsx or .csv first.")
    else:
        raise InventoryAuditError("Choose a .csv or .xlsx inventory file.")

    items = [_row_to_item(row, inventory_uri) for row in rows]
    return inventory_uri, [item for item in items if item.key]


def _read_csv_inventory(path: Path) -> tuple[str, list[dict[str, str]]]:
    inventory_uri = ""
    header: list[str] | None = None
    rows: list[dict[str, str]] = []

    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for raw_row in reader:
            row = [str(cell).strip() for cell in raw_row]
            if not any(row):
                continue
            first = row[0].lower() if row else ""
            if first == "inventory_uri" and len(row) > 1:
                inventory_uri = row[1]
                continue
            if header is None and "key" in {cell.lower() for cell in row}:
                header = [cell.lower() for cell in row]
                continue
            if header:
                padded = row + [""] * max(0, len(header) - len(row))
                rows.append(dict(zip(header, padded)))

    if header is None:
        raise InventoryAuditError("Could not find an inventory header row containing a 'key' column.")
    return inventory_uri, rows


def _read_xlsx_inventory(path: Path) -> tuple[str, list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on app packaging
        raise InventoryAuditError("openpyxl is required for .xlsx input. Install requirements.txt or use CSV.") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    inventory_uri = ""
    header: list[str] | None = None
    rows: list[dict[str, str]] = []

    for raw_row in sheet.iter_rows(values_only=True):
        row = ["" if cell is None else str(cell).strip() for cell in raw_row]
        if not any(row):
            continue
        first = row[0].lower() if row else ""
        if first == "inventory_uri" and len(row) > 1:
            inventory_uri = row[1]
            continue
        if header is None and "key" in {cell.lower() for cell in row}:
            header = [cell.lower() for cell in row]
            continue
        if header:
            padded = row + [""] * max(0, len(header) - len(row))
            rows.append(dict(zip(header, padded)))

    if header is None:
        raise InventoryAuditError("Could not find an inventory header row containing a 'key' column.")
    return inventory_uri, rows


def _row_to_item(row: dict[str, str], inventory_uri: str) -> InventoryItem:
    key = (row.get("key") or row.get("path") or "").strip().lstrip("/")
    s3_uri = (row.get("s3_uri") or row.get("uri") or "").strip()
    bucket = (row.get("bucket") or "").strip()

    if s3_uri and (not bucket or not key):
        parsed_bucket, parsed_key = split_s3_uri(s3_uri)
        bucket = bucket or parsed_bucket
        key = key or parsed_key

    if not bucket and inventory_uri:
        bucket, _prefix = split_s3_uri(inventory_uri)

    if not s3_uri and bucket and key:
        s3_uri = f"s3://{bucket}/{key}"

    size_text = (row.get("size_bytes") or row.get("size") or "0").replace(",", "").strip()
    try:
        size_bytes = int(float(size_text)) if size_text else 0
    except ValueError:
        size_bytes = 0

    return InventoryItem(
        bucket=bucket,
        key=key,
        size_bytes=size_bytes,
        last_modified=(row.get("last_modified") or row.get("lastmodified") or "").strip(),
        s3_uri=s3_uri,
    )


def split_s3_uri(uri: str) -> tuple[str, str]:
    cleaned = uri.strip()
    if not cleaned.startswith("s3://"):
        return "", cleaned.lstrip("/")
    without_scheme = cleaned[5:]
    bucket, _, key = without_scheme.partition("/")
    return bucket, key


def discover_entities(items: Sequence[InventoryItem], inventory_uri: str = "") -> list[AuditEntity]:
    """Discover movie, series, season, and episode audit rows from inventory keys."""

    entities: dict[tuple[str, str], AuditEntity] = {}

    for item in items:
        parts = [part for part in item.key.strip("/").split("/") if part]
        if len(parts) < 2:
            continue

        root = parts[0].lower()
        if root == "movies":
            movie_folder = parts[1]
            title, sku = parse_title_and_sku(movie_folder)
            path_key = f"{parts[0]}/{movie_folder}/"
            entity = _get_entity(entities, "Movie", path_key, title, title, sku, item.bucket)
            entity.files.append(item)
            continue

        if root != "series":
            continue

        series_folder = parts[1]
        series_title, sku = parse_title_and_sku(series_folder)
        series_key = f"{parts[0]}/{series_folder}/"
        series_entity = _get_entity(entities, "Series", series_key, series_title, series_title, sku, item.bucket)

        season_index = _first_matching_index(parts, SEASON_RE, start=2)
        if season_index is None:
            series_entity.files.append(item)
            continue

        season = normalize_season_or_episode(parts[season_index], "s")
        season_key = f"{series_key}{season}/"
        season_name = f"{series_title}_{season}"
        season_entity = _get_entity(entities, "Season", season_key, season_name, series_title, sku, item.bucket, season=season)

        episode_index = _first_matching_index(parts, EPISODE_RE, start=season_index + 1)
        next_part = parts[season_index + 1].lower() if len(parts) > season_index + 1 else ""
        if episode_index is None:
            if next_part not in {"trailer", "trailers"}:
                season_entity.files.append(item)
            continue

        episode = normalize_season_or_episode(parts[episode_index], "e")
        episode_key = f"{season_key}{episode}/"
        episode_name = f"{series_title}_{season}_{episode}"
        episode_entity = _get_entity(
            entities,
            "Episode",
            episode_key,
            episode_name,
            series_title,
            sku,
            item.bucket,
            season=season,
            episode=episode,
        )
        episode_entity.files.append(item)

    if not entities and inventory_uri:
        bucket, prefix = split_s3_uri(inventory_uri)
        if prefix.startswith("movies/"):
            folder = prefix.strip("/").split("/")[1]
            title, sku = parse_title_and_sku(folder)
            _get_entity(entities, "Movie", f"movies/{folder}/", title, title, sku, bucket)
        elif prefix.startswith("series/"):
            folder = prefix.strip("/").split("/")[1]
            title, sku = parse_title_and_sku(folder)
            _get_entity(entities, "Series", f"series/{folder}/", title, title, sku, bucket)

    return sorted(entities.values(), key=_entity_sort_key)


def _get_entity(
    entities: dict[tuple[str, str], AuditEntity],
    content_type: str,
    path_key: str,
    name: str,
    title: str,
    sku: str,
    bucket: str,
    season: str = "",
    episode: str = "",
) -> AuditEntity:
    key = (content_type, path_key)
    if key not in entities:
        entities[key] = AuditEntity(
            content_type=content_type,
            name=name,
            title=title,
            sku=sku,
            bucket=bucket,
            path_key=path_key,
            season=season,
            episode=episode,
        )
    elif not entities[key].bucket and bucket:
        entities[key].bucket = bucket
    return entities[key]


def _first_matching_index(parts: Sequence[str], regex: re.Pattern[str], start: int = 0) -> int | None:
    for index in range(start, len(parts)):
        if regex.match(parts[index]):
            return index
    return None


def _entity_sort_key(entity: AuditEntity) -> tuple[str, int, int, int, str]:
    type_order = {"Movie": 0, "Series": 1, "Season": 2, "Episode": 3}
    season_num = int(entity.season[1:]) if entity.season[1:].isdigit() else 0
    episode_num = int(entity.episode[1:]) if entity.episode[1:].isdigit() else 0
    root_key = "/".join(entity.path_key.strip("/").split("/")[:2])
    return (root_key, type_order.get(entity.content_type, 99), season_num, episode_num, entity.path_key)


def normalize_season_or_episode(value: str, prefix: str) -> str:
    digits = re.sub(r"\D", "", value)
    if not digits:
        return value.lower()
    return f"{prefix}{int(digits):02d}"


def parse_title_and_sku(folder_name: str) -> tuple[str, str]:
    tokens = [token for token in folder_name.strip("_").split("_") if token]
    if not tokens:
        return folder_name, ""

    tail = tokens[-1]
    if _looks_like_trailing_sku(tail):
        title_tokens = tokens[:-1] or [folder_name]
        return "_".join(title_tokens), tail

    head = tokens[0]
    if head.isdigit() and len(head) >= 3 and len(tokens) > 1:
        return "_".join(tokens[1:]), head

    return folder_name, ""


def _looks_like_trailing_sku(token: str) -> bool:
    if UUID_RE.match(token):
        return True
    return token.isdigit() and len(token) >= 6


def build_audit_rows(entities: Iterable[AuditEntity]) -> list[OrderedDict[str, str]]:
    return [audit_entity(entity) for entity in entities]


def audit_entity(entity: AuditEntity) -> OrderedDict[str, str]:
    media_assets = _best_media_assets(entity.files)
    art_assets = _best_art_assets(entity.files)
    media_applicable = entity.content_type in {"Movie", "Episode"}

    values: dict[str, str] = {}
    values["content_type"] = entity.content_type
    values["name"] = entity.name
    values["title"] = entity.title
    values["sku"] = entity.sku
    values["season"] = entity.season
    values["episode"] = entity.episode
    values["s3_path"] = entity.s3_path
    values["file_count"] = str(len(entity.files))

    for field_name, extension in (("mov", ".mov"), ("vtt", ".vtt"), ("srt", ".srt")):
        asset = media_assets.get(extension)
        values[field_name] = "yes" if asset and media_applicable else "n/a" if not media_applicable else "no"
        values[f"{field_name}_file"] = asset.filename if asset else ""

    for art_field in ART_FIELDS:
        asset = art_assets.get(art_field)
        values[art_field] = f"{asset[0]}x{asset[1]}" if asset else ""

    for endpoint in ENDPOINT_ORDER:
        missing = _missing_requirements(values, endpoint, entity.content_type)
        values[endpoint] = "complete" if not missing else "incomplete"
        values[f"missing_{_safe_header(endpoint)}"] = ", ".join(missing)

    return OrderedDict((header, values.get(header, "")) for header in report_headers())


def _best_media_assets(files: Iterable[InventoryItem]) -> dict[str, InventoryItem]:
    best: dict[str, InventoryItem] = {}
    for item in files:
        if item.suffix not in MEDIA_EXTENSIONS:
            continue
        current = best.get(item.suffix)
        if current is None or item.size_bytes > current.size_bytes:
            best[item.suffix] = item
    return best


def _best_art_assets(files: Iterable[InventoryItem]) -> dict[str, tuple[int, int, int, str]]:
    best: dict[str, tuple[int, int, int, str]] = {}
    for item in files:
        if item.suffix not in IMAGE_EXTENSIONS:
            continue
        parsed = parse_art_asset(item.filename)
        if not parsed:
            continue
        field_name, width, height = parsed
        if field_name not in ART_FIELDS:
            continue
        score = (width * height, item.size_bytes)
        current = best.get(field_name)
        if current is None or score > (current[0] * current[1], current[2]):
            best[field_name] = (width, height, item.size_bytes, item.s3_uri)
    return best


def parse_art_asset(filename: str) -> tuple[str, int, int] | None:
    match = ART_RE.search(filename)
    if not match:
        return None
    kind = match.group("kind").lower()
    ratio = match.group("ratio").lower()
    width = int(match.group("width"))
    height = int(match.group("height"))
    return f"{kind}_{ratio}", width, height


def _missing_requirements(row: dict[str, str], endpoint: str, content_type: str) -> list[str]:
    missing: list[str] = []
    for media_field in sorted(get_media_requirements(endpoint, content_type)):
        if row.get(media_field) != "yes":
            missing.append(media_field)
    for art_field in sorted(get_art_requirements(endpoint, content_type)):
        if not row.get(art_field):
            missing.append(art_field)
    return missing


def _safe_header(endpoint: str) -> str:
    return endpoint.replace("+", "plus").replace(" ", "_").replace("-", "_")


def report_headers() -> list[str]:
    """Return the user-facing report column order.

    Endpoint completion columns intentionally come first so the report opens on
    readiness status before lower-level file detail.
    """

    missing_headers = [f"missing_{_safe_header(endpoint)}" for endpoint in ENDPOINT_ORDER]
    identity_headers = ["content_type", "name", "title", "sku", "season", "episode", "s3_path", "file_count"]
    media_headers = ["mov", "mov_file", "vtt", "vtt_file", "srt", "srt_file"]
    return [*ENDPOINT_ORDER, *missing_headers, *identity_headers, *media_headers, *ART_FIELDS]


def write_audit_csv(rows: Sequence[OrderedDict[str, str]], output_path: Path | str) -> Path:
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = report_headers()

    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return destination


def write_audit_xlsx(rows: Sequence[OrderedDict[str, str]], output_path: Path | str) -> Path:
    destination = Path(output_path).with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on app packaging
        raise InventoryAuditError("openpyxl is required to write Excel output. Install requirements.txt.") from exc

    headers = report_headers()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Audit"
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    complete_fill = PatternFill("solid", fgColor="C6EFCE")
    incomplete_fill = PatternFill("solid", fgColor="FFC7CE")
    missing_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    endpoint_cols = {endpoint: headers.index(endpoint) + 1 for endpoint in ENDPOINT_ORDER}
    missing_cols = {f"missing_{_safe_header(endpoint)}": headers.index(f"missing_{_safe_header(endpoint)}") + 1 for endpoint in ENDPOINT_ORDER}
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in endpoint_cols.values():
            cell = worksheet.cell(row_idx, col_idx)
            if cell.value == "complete":
                cell.fill = complete_fill
            elif cell.value == "incomplete":
                cell.fill = incomplete_fill
        for col_idx in missing_cols.values():
            cell = worksheet.cell(row_idx, col_idx)
            if cell.value:
                cell.fill = missing_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, min(worksheet.max_row, 80) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            max_len = max(max_len, len(str(value or "")))
        width = min(max(max_len + 2, 10), 54)
        if header.startswith("missing_") or header.endswith("_file") or header == "s3_path":
            width = min(max(width, 24), 70)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    workbook.save(destination)
    return destination


def resolve_output_paths(output_base: Path | str) -> tuple[Path, Path]:
    base = Path(output_base)
    if base.suffix.lower() == ".csv":
        return base, base.with_suffix(".xlsx")
    if base.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return base.with_suffix(".csv"), base.with_suffix(".xlsx")
    return base.with_suffix(".csv"), base.with_suffix(".xlsx")


def audit_items(items: Sequence[InventoryItem], inventory_uri: str, output_base: Path | str) -> AuditResult:
    entities = discover_entities(items, inventory_uri)
    rows = build_audit_rows(entities)
    csv_path, xlsx_path = resolve_output_paths(output_base)
    write_audit_csv(rows, csv_path)
    write_audit_xlsx(rows, xlsx_path)
    return AuditResult(csv_path=csv_path, xlsx_path=xlsx_path, rows=rows, source_file_count=len(items))


def run_audit(input_path: Path | str, output_path: Path | str) -> AuditResult:
    inventory_uri, items = read_inventory(input_path)
    return audit_items(items, inventory_uri, output_path)


def run_s3_audit(
    s3_uri: str,
    output_path: Path | str,
    progress_callback: Callable[[str], None] | None = None,
) -> AuditResult:
    from inventory_audit_s3 import scan_inventory_to_csv

    inventory_csv_path, inventory_uri, items = scan_inventory_to_csv(
        s3_uri,
        output_path,
        progress_callback=progress_callback,
    )
    result = audit_items(items, inventory_uri, output_path)
    return AuditResult(
        csv_path=result.csv_path,
        xlsx_path=result.xlsx_path,
        rows=result.rows,
        source_file_count=result.source_file_count,
        inventory_csv_path=inventory_csv_path,
    )


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Audit S3 inventory CSV/XLSX files against endpoint requirements.")
    parser.add_argument("input", nargs="?", help="Inventory CSV or XLSX file")
    parser.add_argument("--s3", dest="s3_uri", help="S3 URI to scan directly, such as s3://bucket/movies/")
    parser.add_argument("-o", "--output", help="Output audit base path. Both .csv and .xlsx files are written.")
    args = parser.parse_args(argv)

    if bool(args.input) == bool(args.s3_uri):
        parser.error("Provide exactly one inventory file path or --s3 s3://bucket/prefix.")

    try:
        if args.s3_uri:
            output_path = Path(args.output) if args.output else Path("inventory_audit_s3_report")
            result = run_s3_audit(args.s3_uri, output_path)
        else:
            input_path = Path(str(args.input))
            output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_audit")
            result = run_audit(input_path, output_path)
    except InventoryAuditError as exc:
        print(str(exc))
        return 1
    except Exception as exc:  # noqa: BLE001
        if exc.__class__.__name__ == "InventoryAuditError":
            print(str(exc))
        else:
            print(f"Unexpected error: {exc}")
        return 1

    print(f"Wrote {result.csv_path}")
    print(f"Wrote {result.xlsx_path}")
    if result.inventory_csv_path:
        print(f"Wrote {result.inventory_csv_path}")
    print(f"Audited {result.entity_count} row(s) from {result.source_file_count} inventory file(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
