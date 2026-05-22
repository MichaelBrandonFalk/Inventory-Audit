"""Tkinter desktop app for Inventory Audit."""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import re
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from inventory_audit_core import InventoryAuditError, run_audit
from inventory_audit_requirements import (
    APP_NAME,
    APP_VERSION,
    CONTENT_TYPES,
    CUSTOM_REQUIREMENT_FIELD_OPTIONS,
    ENDPOINT_ORDER,
    OPTIONAL_ART_FIELDS,
    get_art_requirements,
    get_media_requirements,
    is_endpoint_applicable,
)
from inventory_audit_s3 import scan_inventory_to_csv


APP_TITLE = f"{APP_NAME} v{APP_VERSION}"


class InventoryAuditApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("940x680")

        self.input_path = tk.StringVar()
        self.s3_uri = tk.StringVar()
        self.output_dir = tk.StringVar(value=str(_default_output_dir()))
        self.status_text = tk.StringVar(value="Ready")
        self.running = False
        self.run_started_at: float | None = None
        self.current_phase = "Ready"
        self.heartbeat_after_id: str | None = None
        self.custom_requirements: dict[str, set[str]] = {content_type: set() for content_type in CONTENT_TYPES}

        self._build()

    def _build(self) -> None:
        frame = ttk.Frame(self.root, padding=18)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text=APP_TITLE, font=("Helvetica", 18, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(
            frame,
            text="Load an S3 inventory CSV/XLSX or scan an S3 path directly, then export CSV and Excel endpoint readiness reports.",
            wraplength=840,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 18))

        self._file_row(frame, 2, "Inventory File", self.input_path, self._browse_input)
        self._entry_row(frame, 3, "S3 Path", self.s3_uri, "Optional direct scan, e.g. s3://bucket/movies/")
        self._file_row(frame, 4, "Output Folder", self.output_dir, self._browse_output, directory=True)

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=5, column=0, columnspan=3, sticky="w", pady=(12, 18))
        self.file_button = ttk.Button(button_frame, text="Audit File", command=self._run_file)
        self.file_button.pack(side="left", padx=(0, 10))
        self.inventory_button = ttk.Button(button_frame, text="Create S3 Inventory CSV", command=self._run_s3_inventory)
        self.inventory_button.pack(side="left", padx=(0, 10))
        self.s3_button = ttk.Button(button_frame, text="Audit S3 Path", command=self._run_s3)
        self.s3_button.pack(side="left", padx=(0, 10))
        self.requirements_button = ttk.Button(button_frame, text="Audit Requirements", command=self._open_requirements_window)
        self.requirements_button.pack(side="left")

        ttk.Label(frame, textvariable=self.status_text).grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 8))

        self.log = scrolledtext.ScrolledText(frame, height=28, wrap="word")
        self.log.grid(row=7, column=0, columnspan=3, sticky="nsew")
        frame.rowconfigure(7, weight=1)

        self._log("Ready. Choose an inventory export or enter an S3 path to start.")

    def _file_row(self, frame: ttk.Frame, row: int, label: str, variable: tk.StringVar, callback, directory: bool = False) -> None:
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(12, 12), pady=5)
        button_label = "Browse Folder" if directory else "Browse File"
        ttk.Button(frame, text=button_label, command=callback).grid(row=row, column=2, sticky="ew", pady=5)

    def _entry_row(self, frame: ttk.Frame, row: int, label: str, variable: tk.StringVar, hint: str = "") -> None:
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(12, 12), pady=5)
        ttk.Label(frame, text=hint).grid(row=row, column=2, sticky="w", pady=5)

    def _browse_input(self) -> None:
        selected = filedialog.askopenfilename(filetypes=[("Inventory Files", "*.csv *.xlsx *.xlsm"), ("All Files", "*.*")])
        if selected:
            self.input_path.set(selected)
            input_path = Path(selected)
            self._log(f"Selected {input_path}")

    def _browse_output(self) -> None:
        selected = filedialog.askdirectory()
        if selected:
            self.output_dir.set(selected)

    def _run_file(self) -> None:
        input_value = self.input_path.get().strip()
        output_value = self.output_dir.get().strip()
        if not input_value:
            messagebox.showerror(APP_TITLE, "Choose an inventory CSV or XLSX file.")
            return
        if not output_value:
            messagebox.showerror(APP_TITLE, "Choose an output folder.")
            return

        input_path = Path(input_value)
        output_path = Path(output_value) / f"{input_path.stem}_inventory_audit_v{APP_VERSION.replace('.', '_')}"

        self._log(f"Running audit for {input_path} ...")
        try:
            result = run_audit(input_path, output_path, self._custom_requirements_for_run())
        except (InventoryAuditError, OSError) as exc:
            self._log(f"Audit failed: {exc}")
            messagebox.showerror(APP_TITLE, str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            self._log(f"Unexpected error: {exc}")
            messagebox.showerror(APP_TITLE, f"Unexpected error: {exc}")
            return

        self._handle_result(result)

    def _run_s3(self) -> None:
        self._start_s3_run(audit_after_inventory=True)

    def _run_s3_inventory(self) -> None:
        self._start_s3_run(audit_after_inventory=False)

    def _start_s3_run(self, audit_after_inventory: bool) -> None:
        if self.running:
            return
        s3_value = self.s3_uri.get().strip()
        output_value = self.output_dir.get().strip()
        if not s3_value:
            messagebox.showerror(APP_TITLE, "Enter an S3 path like s3://bucket/path/.")
            return
        if not output_value:
            messagebox.showerror(APP_TITLE, "Choose an output folder.")
            return

        output_path = Path(output_value) / f"{_safe_output_stem(s3_value)}_inventory_audit_v{APP_VERSION.replace('.', '_')}"
        self._set_running(True)
        mode = "S3 audit" if audit_after_inventory else "S3 inventory export"
        self._record_progress(f"Starting {mode} for {s3_value}")
        self._log("This is read-only. The app will list S3 objects and write local report files only.")

        worker = threading.Thread(
            target=self._run_s3_worker,
            args=(s3_value, output_path, audit_after_inventory),
            daemon=True,
        )
        worker.start()

    def _run_s3_worker(self, s3_uri: str, output_path: Path, audit_after_inventory: bool) -> None:
        try:
            self._thread_progress("Worker started.")
            inventory_csv_path, _inventory_uri, items = scan_inventory_to_csv(
                s3_uri,
                output_path,
                progress_callback=self._thread_progress,
            )
            if audit_after_inventory:
                self._thread_progress(f"Step 2/2: auditing raw inventory CSV {inventory_csv_path}")
                result = run_audit(inventory_csv_path, output_path, self._custom_requirements_for_run())
                result = replace(result, inventory_csv_path=inventory_csv_path)
                self._thread_progress("Audit report generation complete.")
            else:
                self.root.after(0, self._handle_inventory_result, inventory_csv_path, len(items))
        except (InventoryAuditError, OSError) as exc:
            self.root.after(0, self._handle_error, str(exc))
        except Exception as exc:  # noqa: BLE001
            self.root.after(0, self._handle_error, f"Unexpected error: {exc}")
        else:
            if audit_after_inventory:
                self.root.after(0, self._handle_result, result)
        finally:
            self.root.after(0, self._set_running, False)

    def _thread_progress(self, message: str) -> None:
        self.root.after(0, self._record_progress, message)

    def _record_progress(self, message: str) -> None:
        self.current_phase = message
        self.status_text.set(message)
        self._log(f"S3: {message}")

    def _handle_result(self, result) -> None:
        if getattr(result, "inventory_csv_path", None):
            self._log(f"Step 1 complete. Raw inventory CSV: {result.inventory_csv_path}")
        self._log(f"Step 2 complete. Wrote audit CSV: {result.csv_path}")
        self._log(f"Step 2 complete. Wrote audit XLSX: {result.xlsx_path}")
        self._log(f"Audited {result.entity_count} row(s) from {result.source_file_count} inventory file(s).")
        self._log(_summarize_endpoint_statuses(result.rows))
        self.status_text.set(f"Complete. Audited {result.entity_count} row(s).")
        messagebox.showinfo(APP_TITLE, f"Wrote audit reports:\n{result.csv_path}\n{result.xlsx_path}")

    def _handle_inventory_result(self, inventory_csv_path: Path, object_count: int) -> None:
        self._log(f"Inventory export complete. Listed {object_count} object(s).")
        self._log(f"Raw inventory CSV: {inventory_csv_path}")
        self.status_text.set(f"Inventory export complete. Listed {object_count} object(s).")
        messagebox.showinfo(APP_TITLE, f"Inventory CSV created:\n{inventory_csv_path}\n\nObjects listed: {object_count}")

    def _handle_error(self, message: str) -> None:
        self._log(f"Audit failed: {message}")
        self.status_text.set("Failed")
        messagebox.showerror(APP_TITLE, message)

    def _set_running(self, running: bool) -> None:
        self.running = running
        if running:
            self.run_started_at = time.monotonic()
            self._schedule_heartbeat()
        else:
            self.run_started_at = None
            if self.heartbeat_after_id is not None:
                self.root.after_cancel(self.heartbeat_after_id)
                self.heartbeat_after_id = None
        state = "disabled" if running else "normal"
        self.file_button.configure(state=state)
        self.inventory_button.configure(state=state)
        self.s3_button.configure(state=state)
        self.requirements_button.configure(state=state)

    def _schedule_heartbeat(self) -> None:
        if self.heartbeat_after_id is not None:
            self.root.after_cancel(self.heartbeat_after_id)
        self.heartbeat_after_id = self.root.after(30000, self._heartbeat)

    def _heartbeat(self) -> None:
        self.heartbeat_after_id = None
        if not self.running or self.run_started_at is None:
            return
        elapsed_seconds = int(time.monotonic() - self.run_started_at)
        minutes, seconds = divmod(elapsed_seconds, 60)
        self._log(f"Still working ({minutes}m {seconds:02d}s): {self.current_phase}")
        self._schedule_heartbeat()

    def _log(self, message: str) -> None:
        self.log.insert("end", f"{message}\n")
        self.log.see("end")

    def _custom_requirements_for_run(self) -> dict[str, set[str]]:
        return {content_type: set(fields) for content_type, fields in self.custom_requirements.items() if fields}

    def _open_requirements_window(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("Audit Requirements")
        window.geometry("900x520")
        window.columnconfigure(0, weight=1)
        window.rowconfigure(0, weight=1)

        columns = ("endpoint", "content_type", "status", "required", "optional")
        tree = ttk.Treeview(window, columns=columns, show="headings")
        headings = {
            "endpoint": "Endpoint",
            "content_type": "Type",
            "status": "Status",
            "required": "Required Fields",
            "optional": "Optional/Tracked",
        }
        widths = {"endpoint": 120, "content_type": 90, "status": 90, "required": 460, "optional": 140}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="w")
        tree.grid(row=0, column=0, sticky="nsew", padx=12, pady=(12, 8))

        scrollbar = ttk.Scrollbar(window, orient="vertical", command=tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=(12, 8))
        tree.configure(yscrollcommand=scrollbar.set)

        for row in self._requirements_rows():
            tree.insert("", "end", values=row)

        button_frame = ttk.Frame(window)
        button_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=12, pady=(0, 12))
        ttk.Button(button_frame, text="Export XLSX", command=lambda: self._export_requirements_xlsx(window)).pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Add Custom Audit Requirements", command=self._open_custom_requirements_window).pack(side="left")

    def _requirements_rows(self) -> list[tuple[str, str, str, str, str]]:
        rows: list[tuple[str, str, str, str, str]] = []
        for endpoint in ENDPOINT_ORDER:
            for content_type in CONTENT_TYPES:
                if not is_endpoint_applicable(endpoint, content_type):
                    rows.append((endpoint, content_type, "N/A", "N/A", ""))
                    continue
                required = sorted(get_media_requirements(endpoint, content_type) | get_art_requirements(endpoint, content_type))
                rows.append((endpoint, content_type, "Required" if required else "No required fields", ", ".join(required), ", ".join(sorted(OPTIONAL_ART_FIELDS))))

        for content_type in CONTENT_TYPES:
            fields = sorted(self.custom_requirements.get(content_type, set()))
            rows.append(("custom audit", content_type, "Required" if fields else "N/A", ", ".join(fields) if fields else "N/A", ""))
        return rows

    def _export_requirements_xlsx(self, parent: tk.Toplevel) -> None:
        selected = filedialog.asksaveasfilename(
            parent=parent,
            title="Export Audit Requirements",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"Inventory_Audit_requirements_v{APP_VERSION.replace('.', '_')}.xlsx",
        )
        if not selected:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            messagebox.showerror(APP_TITLE, f"openpyxl is required to export XLSX files: {exc}", parent=parent)
            return

        headers = ["Endpoint", "Type", "Status", "Required Fields", "Optional/Tracked"]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Audit Requirements"
        worksheet.append(headers)
        for row in self._requirements_rows():
            worksheet.append(list(row))

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for row_idx in range(2, worksheet.max_row + 1):
                max_len = max(max_len, len(str(worksheet.cell(row_idx, col_idx).value or "")))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)
        workbook.save(selected)
        messagebox.showinfo(APP_TITLE, f"Exported audit requirements:\n{selected}", parent=parent)

    def _open_custom_requirements_window(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("Custom Audit Requirements")
        window.geometry("560x520")

        type_frame = ttk.LabelFrame(window, text="Apply To")
        type_frame.pack(fill="x", padx=12, pady=(12, 8))
        field_frame = ttk.LabelFrame(window, text="Required Fields")
        field_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        existing = self._custom_requirements_for_run()
        existing_types = {content_type for content_type, fields in existing.items() if fields}
        existing_fields = {field for fields in existing.values() for field in fields}

        type_vars: dict[str, tk.BooleanVar] = {}
        for idx, content_type in enumerate(CONTENT_TYPES):
            variable = tk.BooleanVar(value=content_type in existing_types if existing_types else content_type in {"Movie", "Episode"})
            type_vars[content_type] = variable
            ttk.Checkbutton(type_frame, text=content_type, variable=variable).grid(row=0, column=idx, sticky="w", padx=8, pady=8)

        field_vars: dict[str, tk.BooleanVar] = {}
        for idx, field_name in enumerate(CUSTOM_REQUIREMENT_FIELD_OPTIONS):
            variable = tk.BooleanVar(value=field_name in existing_fields)
            field_vars[field_name] = variable
            ttk.Checkbutton(field_frame, text=field_name, variable=variable).grid(row=idx // 3, column=idx % 3, sticky="w", padx=10, pady=5)

        button_frame = ttk.Frame(window)
        button_frame.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(button_frame, text="Save Custom Requirements", command=lambda: self._save_custom_requirements(window, type_vars, field_vars)).pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Clear Custom Requirements", command=lambda: self._clear_custom_requirements(window)).pack(side="left")

    def _save_custom_requirements(
        self,
        window: tk.Toplevel,
        type_vars: dict[str, tk.BooleanVar],
        field_vars: dict[str, tk.BooleanVar],
    ) -> None:
        selected_types = [content_type for content_type, variable in type_vars.items() if variable.get()]
        selected_fields = {field_name for field_name, variable in field_vars.items() if variable.get()}
        self.custom_requirements = {content_type: set() for content_type in CONTENT_TYPES}
        for content_type in selected_types:
            self.custom_requirements[content_type] = set(selected_fields)
        self._log(f"Custom audit requirements updated: {self._custom_requirement_summary()}")
        window.destroy()

    def _clear_custom_requirements(self, window: tk.Toplevel) -> None:
        self.custom_requirements = {content_type: set() for content_type in CONTENT_TYPES}
        self._log("Custom audit requirements cleared.")
        window.destroy()

    def _custom_requirement_summary(self) -> str:
        parts = []
        for content_type in CONTENT_TYPES:
            fields = sorted(self.custom_requirements.get(content_type, set()))
            if fields:
                parts.append(f"{content_type}={','.join(fields)}")
        return "; ".join(parts) if parts else "none"


def _summarize_endpoint_statuses(rows: list[dict[str, str]]) -> str:
    if not rows:
        return "No audit rows were detected."

    endpoints = ["Axinom", "Amazon", "Roku", "Frndly", "T+", "YouTube"]
    lines = ["Endpoint completion summary:"]
    for endpoint in endpoints:
        complete = sum(1 for row in rows if row.get(endpoint) == "complete")
        lines.append(f"- {endpoint}: {complete}/{len(rows)} complete")
    return "\n".join(lines)


def _safe_output_stem(value: str) -> str:
    cleaned = value.replace("s3://", "").strip("/").replace("/", "_")
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", cleaned).strip("_")
    return cleaned or "s3_inventory"


def _default_output_dir() -> Path:
    downloads = Path.home() / "Downloads"
    return downloads if downloads.exists() else Path.home()


def main() -> None:
    root = tk.Tk()
    InventoryAuditApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
